import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from datetime import datetime

'''Ordem de colunas na tabela é fundamental.
A ordem para o funcionamento do programa é:
Vídeo, Título, Data, Likes, Visualizações, impressões'''

# definir ano e mes atual
currentMonth = datetime.now().month
currentYear = datetime.now().year


# carregar tabela do youtube
yt_data = openpyxl.load_workbook('Dados_FDP_Fevereiro.xlsx')
yt_sheet = yt_data.active
# carregar tabela da wd
wd_data = openpyxl.load_workbook('Formulário WD-40.xlsx')
wd_sheet = wd_data['Abril']


# quantifica numero de linhas e colunas na tabela
yt_rows = yt_sheet.max_row
yt_columns = yt_sheet.max_column
yt_last_cell = str(openpyxl.utils.get_column_letter(yt_columns) + str(yt_rows))


# lista para armazenar todos os vídeos do periodo
videos_list = []


# loop seleciona apenas os vídeos do mes e insere na lista
for row in tuple(yt_sheet['A3':yt_last_cell]):
    if row[2].value.month == currentMonth - 1 and row[2].value.year == currentYear:
        videos_list.append(list(x.value for x in row))
    else:
        continue


# Loop insere valores das listas nas células.
counter = 0
for video in videos_list:
    wd_sheet['A' + str(4 + counter)] = 'Youtube'
    wd_sheet['B' + str(4 + counter)] = '@foradaspistas'
    wd_sheet['C' + str(4 + counter)] = 'Vídeo'
    wd_sheet['D' + str(4 + counter)] = 'EVD'
    wd_sheet['E' + str(4 + counter)].hyperlink = 'https://youtube.com/watch?v=' + video[0]
    wd_sheet['F' + str(4 + counter)] = video[2].strftime('%d/%m/%y')
    wd_sheet['G' + str(4 + counter)] = int(video[5])
    wd_sheet['H' + str(4 + counter)] = None
    wd_sheet['I' + str(4 + counter)] = int(video[3])
    counter += 1


# Salvar cópia .xlsx
wd_data.save('FormulárioWD_copycopy.xlsx')
