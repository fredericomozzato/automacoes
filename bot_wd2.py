import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from datetime import datetime
import sys
import os

'''Ordem de colunas na tabela é fundamental.
A ordem para o funcionamento do programa é:
Vídeo, Título, Data, Likes, Visualizações, impressões'''

tabelaYt = sys.argv[1]
mes = sys.argv[2]

# referencia para meses:
monthRefer = {
    '1': 'Janeiro',
    '2': 'Fevereiro',
    '3': 'Março',
    '4': 'Abril',
    '5': 'Maio',
    '6': 'Junho',
    '7': 'Julho',
    '8': 'Agosto',
    '9': 'Setembro',
    '10': 'Outubro',
    '11': 'Novembro',
    '12': 'Dezembro'
}

# definir mes a ano
currentMonth = datetime.now().month
currentYear = datetime.now().year

# carregar tabela de dados do Youtube
ytData = openpyxl.load_workbook(tabelaYt)
ytSheet = ytData.active

# carregar tablea WD
tabelaWd = openpyxl.load_workbook('Formulário WD-40.xlsx')
wdSheet = tabelaWd[monthRefer[mes]]


# quantificar numero de linhas e colunas
ytRows = ytSheet.max_row
ytCol = ytSheet.max_column
ytLastCell = str(openpyxl.utils.get_column_letter(ytCol) + str(ytRows))

# lista com videos do periodo desejado
videoList = []

# loop para selecionar videos do periodo
for row in tuple(ytSheet['A3':ytLastCell]):
    if row[2].value is None:
        continue
    if row[2].value.month == currentMonth and row[2].value.year == currentYear:
        videoList.append(list(cell.value for cell in row))
    else:
        continue

# ordena lista de videos por data em ordem descendente
sortList = sorted(videoList, key=lambda x: datetime.strftime(x[2], '%d/%m/%y'))

# loop para inserir valores na tabela
counter = 0
for video in sortList:
    wdSheet['A' + str(4 + counter)] = 'Youtube'
    wdSheet['B' + str(4 + counter)] = '@foradaspistas'
    wdSheet['C' + str(4 + counter)] = 'Vídeo'
    wdSheet['D' + str(4 + counter)] = 'EVD'
    wdSheet['E' + str(4 + counter)].hyperlink = 'https://youtube.com/watch?v=' + video[0]
    wdSheet['F' + str(4 + counter)] = video[2].strftime('%d/%m/%y')
    wdSheet['G' + str(4 + counter)] = int(video[5])
    wdSheet['H' + str(4 + counter)] = None
    wdSheet['I' + str(4 + counter)] = int(video[3])
    counter += 1

# salva uma copia da tabela atualizada
tabelaWd.save(f'Formulário WD-40_{monthRefer[mes]}.xlsx')
